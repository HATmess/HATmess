from openpyxl import load_workbook
import PySimpleGUI as sg
from pathlib import Path
import ipaddress
import time
import os
import sys
import re

def excel_to_dict():
        
    sheet = book[sheets[1]]
    device = {}
    row = 4

    device = {
        'hostname'        : str(sheet.cell(row,2).value).rsplit('-',1)[0] ,
        'snmp_location'   : str(sheet.cell(row,3).value),
        'ASN'             : str(sheet.cell(row,4).value) ,
        'lo0_mgt'         : str(sheet.cell(row,5).value) ,
        'lo0_proxy'       : str(sheet.cell(row,6).value) ,
        'lo0_bgp'         : str(sheet.cell(row,7).value) ,
        'pe1_vrf_alger'   : str(sheet.cell(row,8).value) ,
        'pe1_vrf_annaba'  : str(sheet.cell(row,9).value) ,
        'pe1_vrf_oran'    : str(sheet.cell(row,10).value) ,
        'pe2_vrf_alger'   : str(sheet.cell(row,11).value) ,
        'pe2_vrf_annaba'  : str(sheet.cell(row,12).value) ,
        'pe2_vrf_oran'    : str(sheet.cell(row,13).value) ,
        'pe1_global'      : str(sheet.cell(row,14).value) ,
        'pe2_global'      : str(sheet.cell(row,15).value) ,
        'twamp_address'   : str(sheet.cell(row,16).value) ,
        'nas_id'          : str(sheet.cell(row,17).value) ,
        'pe1_bgp_alger'   : str(sheet.cell(row,18).value) ,
        'pe1_bgp_annaba'  : str(sheet.cell(row,19).value) ,
        'pe1_bgp_oran'    : str(sheet.cell(row,20).value) ,
        'pe2_bgp_alger'   : str(sheet.cell(row,21).value) ,
        'pe2_bgp_annaba'  : str(sheet.cell(row,22).value) ,
        'pe2_bgp_oran'    : str(sheet.cell(row,23).value) ,
        'ospf_interco'    : str(sheet.cell(row,24).value) ,
        'neighbor_bgp_pe1': str(sheet.cell(row,25).value) ,
        'neighbor_bgp_pe2': str(sheet.cell(row,26).value) 
        }
        
    return device

 

def get_ip (var , inc) :
    return str(ipaddress.ip_address(device.get(var))+inc)

def port_mapping(port):
    port_map = {
        "card1/pfe1": "et-0/0/5",
        "card1/pfe2": "et-0/1/5",
        "card2/pfe1": "et-8/0/5",
        "card2/pfe2": "et-8/1/5"
    }
    return port_map[port]


def load_basic_config(dev):

    #dict = {'ASN': 65119, 'hostname': "CTP-STF-002", 'lo0_mgt': '10.104.119.101', 'ospf_interco': '10.104.19.250', 'pe1_vrf_alger': '10.104.119.2', 'pe2_vrf_alger': '10.104.119.14', 'pe1_global': '10.210.19.242','locatin': "CTP-BNG SETIF", 'twamp_address': '10.201.119.1', 'nas_id': "J_STF2", 'pe1_bgp_alger': '10.19.19.16', 'pe1_bgp_annaba': '10.19.19.23', 'pe1_bgp_oran': '10.19.19.31'}
    
    #dict2 = {'lo0_proxy': get_ip('lo0_mgt', 1), 'lo0_bgp': get_ip('lo0_mgt', 4), 'pe1_vrf_annaba': get_ip('pe1_vrf_alger' ,4), 'pe1_vrf_oran': get_ip('pe1_vrf_alger' , 8), 'pe2_vrf_annaba' : get_ip('pe2_vrf_alger' , 4), 'pe2_vrf_oran': get_ip('pe2_vrf_alger', 8), 'neighbor_bgp_pe1' : get_ip ('pe1_global',-1) ,'neighbor_bgp_pe2' : get_ip ('pe1_global', 3) , 'pe2_global': get_ip('pe1_global' , 4),  'pe2_bgp_alger': get_ip ('pe1_bgp_alger' , 25600), 'pe2_bgp_annaba': get_ip('pe1_bgp_annaba' ,25600), 'pe2_bgp_oran': get_ip('pe1_bgp_oran',25600)}
    
    basic_conf = open("b_con", "rt")
    data = basic_conf.read()
    
    #input = { **dict , **dict2 }
    
    for key,value in dev.items():
        data = data.replace(key,str(value))
    
    basic_conf.close()
    conf_file = str("..\\BNG_Generated_Conf\\Basic-" + dev['hostname']) + "_BNG.conf"
    basic_conf = open(conf_file , "wt")
    basic_conf.write(data)
    basic_conf.close()
    
    return (conf_file)

def svlan_profile ():
    sheet = book[sheets[2]]
    data = ['### Interfaces \n\n']
    lldp = ['\n\n']
    port_list = []
    for row in range (2 ,sheet.max_row ):
        port = str(sheet.cell(row , column = 8).value).strip()
        svlan = re.findall(r'\d+', str (sheet.cell(row, column = 4).value))[0]
        cvlan = str (sheet.cell(row , column = 3).value).split(',')
        pado = 'delay' in str (sheet.cell(row , column = 4).value)

        if 'et-' not in port :
            port = port_mapping (port.lower())

        if not port in port_list:
            port_list.append (port)
            data.append(f"\n### {port} \n\n")
            data.append(f"set interfaces {port} description {device['hostname']}-BNG01_TO_{device['hostname']}-ASBR01_{port}_100G\n")
            data.append(f"set interfaces {port} traps \n")
            data.append(f"set interfaces {port} flexible-vlan-tagging\n")
            data.append(f"set interfaces {port} auto-configure remove-when-no-subscribers\n")
            data.append(f"set interfaces {port} mtu 9192\n")
            data.append(f"set interfaces {port} hold-time up 5000\n")
            data.append(f"set interfaces {port} hold-time down 0\n")
            data.append(f"set interfaces {port} encapsulation flexible-ethernet-services\n")
            data.append(f"set interfaces {port} gigether-options ignore-l3-incompletes\n")
            data.append(f"set interfaces {port} auto-configure stacked-vlan-ranges dynamic-profile SVLAN_PROFILE accept pppoe\n")

            if len(port_list) == 1 and selected_interface != "" :
                lldp.append(f"set interfaces {selected_interface} unit 4021 vlan-id 4021\n")
                lldp.append(f"set interfaces {selected_interface} unit 4021 family inet mtu 1500\n")
                lldp.append(f"set interfaces {selected_interface} unit 4021 family inet address {device['ospf_interco']}/30\n")
                lldp.append(f"set routing-instances RESIDENTIAL protocols ospf area 0.0.0.0 interface {port}.4021\n")
                lldp.append(f"set routing-instances RESIDENTIAL interface {selected_interface}.4021\n")
            lldp.append(f"set protocols lldp interface {port}\n")
        if not pado :
            for value in cvlan:
                data.append(f"set interfaces {port} auto-configure stacked-vlan-ranges dynamic-profile SVLAN_PROFILE ranges {svlan}-{svlan},{value}\n")
        else:
            for value in cvlan:
                data.append(f"set interfaces {port} auto-configure stacked-vlan-ranges dynamic-profile SVLAN_PROFILE_DELAYED ranges {svlan}-{svlan},{value}\n")
       
    with open("..\\BNG_Generated_Conf\\interfaces.conf", "w", encoding='utf-8') as intf :
        intf.write(''.join(data))
        intf.write(''.join(lldp))
    intf.close()
    
def static_users_config():
    sheet = book[sheets[3]]
    data = ['\n### STATIC USER ROUTES   \n\n']
    annotation =  ['\n### Routes Descriptions\n\nedit routing-instances jrp routing-options static\n']
    
    for row in range (2 ,sheet.max_row+1 ):
        route = str(sheet.cell(row , column = 3).value).strip()
        next_hop = str(sheet.cell(row , column = 4).value).strip()
        description = str(sheet.cell(row , column = 5).value).strip().upper()
        
        data.append(f"set routing-instances jrp routing-options static route {route} next-hop {next_hop}\n")
        if description != "NONE":
            annotation.append(f"annotate route {route} \"{description}\"\n")
            
    with open("..\\BNG_Generated_Conf\\static_users.conf", "w", encoding='utf-8') as file :
        file.write(''.join(data))
        file.write(''.join(annotation))
    file.close()
  
def policy_config():
    sheet = book[sheets[4]]
    ss = "set routing-instances RESIDENTIAL routing-options static route "
    exp = "set policy-options policy-statement SUBSCRIBERS_EXPORT"
    data = ['## PUBLIC POOLS & POLICIES \n## PUBLIC POOLS\n\n']
    sub_export_t1 = ["\n## SUBSCRIBERS_EXPORT \n\n%s term 1 from protocol static \n"%exp]
    sub_export_t2 = ["\n"]
    access_route = ['\n## ACCESS INTERNAL ROUTES \n\n']
    vrfs = ["\n"]
    static = ["\n\n## STATIC ROUTES RESIDENTIAL \n\n"]
   
    rank = 1
    
    for row in range (2 ,sheet.max_row ):
        pool = str(sheet.cell(row , column = 1).value).strip()
        if pool != 'None' :
            data.append(f"set routing-instances RESIDENTIAL access address-assignment pool V4_PUBLIC_POOL{rank} link V4_PUBLIC_POOL{rank+1}\n")
            data.append(f"set routing-instances RESIDENTIAL access address-assignment pool V4_PUBLIC_POOL{rank} family inet network {pool}\n")
            rank+=1
            sub_export_t1.append(f"{ exp } term 1 from route-filter {pool} exact\n")
            sub_export_t2.append(f"{ exp } term 2 from route-filter {pool} longer\n")
            static.append(f"{ss} {pool} discard \n")

    for row in range (2 , sheet.max_row):
        access = str(sheet.cell(row , column = 2).value).strip()
        if access != 'None' :
            access_route.append(f"set policy-options policy-statement OSPF_REDIS_ACCESS_INTERNAL term 1 from route-filter {access} longer\n")
            sub_export_t1.append(f"{ exp } term 1 from route-filter {access} exact\n")
            sub_export_t2.append(f"{ exp } term 2 from route-filter {access} longer\n")
            static.append(f"{ss} {access} discard \n")
    
    vrfs.append(f"{ss} { device['pe1_bgp_alger'] }/32 next-hop { get_ip('pe1_vrf_alger' ,-1) }\n")
    vrfs.append(f"{ss} { device['pe1_bgp_annaba'] }/32 next-hop { get_ip('pe1_vrf_annaba' ,-1) }\n")
    vrfs.append(f"{ss} { device['pe1_bgp_oran'] }/32 next-hop { get_ip('pe1_vrf_oran' ,-1) }\n")
    vrfs.append(f"{ss} { device['pe2_bgp_alger'] }/32 next-hop { get_ip('pe2_vrf_alger' ,-1) }\n")
    vrfs.append(f"{ss} { device['pe2_bgp_annaba'] }/32 next-hop { get_ip('pe2_vrf_annaba' ,-1) }\n")
    vrfs.append(f"{ss} { device['pe2_bgp_oran'] }/32 next-hop { get_ip('pe2_vrf_oran' ,-1) }\n")

    with open("..\\BNG_Generated_Conf\\policies.conf", "w", encoding='utf-8') as file :
        file.write(''.join(data))
        file.write(''.join(access_route))
        file.write(''.join(sub_export_t1))
        file.write(''.join(sub_export_t2))
        file.write(''.join(static))
        file.write(''.join(vrfs))
    file.close()
    
if __name__ == "__main__":

    print("Loading Input DATA ...")
    lbox = ['et-0/0/5','et-0/1/5','et-8/0/5','et-8/1/5']
    working_directory = os.getcwd()

    layout = [  
            [sg.Text("Choose a data input file:")],
            [sg.InputText(key="-INPUT-"),
            sg.FileBrowse(initial_folder=working_directory, file_types=[("Excel Files", "*.xlsx")])],
            [sg.Listbox(values=lbox , size = (15, 5), key = '_SELECTED_ITEM_', enable_events = True) , sg.Checkbox('OSPF', default=True)],
            [sg.Button('Submit'), sg.Exit()]
        ]

    window = sg.Window("Loading Input File", layout)
    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, 'Exit'):
            sys.exit()
        elif event == "Submit":
            filepath = values['-INPUT-']
            if Path(filepath).is_file():
                try:
                    book = load_workbook(filepath, data_only=True)
                    selected_interface = ''.join(window["_SELECTED_ITEM_"].get())
                    print(f'selected ospf interface = {selected_interface}')
                    break
                except Exception as e:
                    print("Error: ", e)

    start_time = time.time()
    
    sheets = book.sheetnames

    window.close()
    
    
    if os.path.exists("..\\BNG_Generated_Conf"):
        sg.PopupError("Directory already exists ! ", title = "Error !", keep_on_top=True)
        exit()
    else:
        os.mkdir("..\\BNG_Generated_Conf")

    device = excel_to_dict ()   
    generated_conf = load_basic_config(device)
    #print ("Basic configuration have been created successfully..", generated_conf)
    #os.startfile(generated_conf)
    svlan_profile()
    print("Interfaces configuration created.. ",)
    #os.startfile("interfaces.conf")
    static_users_config()
    print("Static users configuration created.. ")
    #os.startfile("static_users.conf")
    policy_config()
    print("Policie\'s configuration created.. ")
    #os.startfile("policies.conf")
    os.startfile("..\\BNG_Generated_Conf")
    
    sg.popup("Success! \nAll configurations have been generated successfully ", title = "Success !",  keep_on_top=True)
    print("Execution Time : \n--- %s seconds ---" % (time.time() - start_time))
